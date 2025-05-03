from pykrx import stock
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime, timedelta
import requests
from bs4 import BeautifulSoup

file_path = "stock.xlsx"
wb = load_workbook(file_path)
ws = wb.active

today = datetime.today()

def to_yyyymmdd(date):
    return date.strftime("%Y%m%d")

def get_last_trading_day(start_date, ticker):
    date = start_date
    while True:
        date_str = to_yyyymmdd(date)
        try:
            df = stock.get_market_ohlcv(date_str, date_str, ticker)
            if not df.empty:
                return date_str
        except:
            pass
        date -= timedelta(days=1)

def format_market_cap(market_cap_won):
    market_cap_억 = market_cap_won // 100_000_000
    if market_cap_억 >= 10_000:
        market_cap_조 = market_cap_억 / 10_000
        return f"{market_cap_조:.2f}조"
    else:
        return f"{market_cap_억}억"

def get_financials_2024(ticker):
    url = f"https://finance.naver.com/item/main.nhn?code={ticker}"
    headers = {'User-Agent': 'Mozilla/5.0'}
    res = requests.get(url, headers=headers)
    soup = BeautifulSoup(res.text, 'html.parser')

    table = soup.find('table', {'class': 'tb_type1 tb_num tb_type1_ifrs'})
    if not table:
        print(f"{ticker}: 재무제표 테이블을 찾을 수 없습니다.")
        return {}

    rows = table.find_all('tr')
    result = {}

    for row in rows:
        th = row.find('th')
        if th:
            row_name = th.text.strip()
            tds = row.find_all('td')
            values = [td.get_text(strip=True).replace(",", "") for td in tds]

            if len(values) >= 3:
                if "매출액" in row_name and "률" not in row_name:
                    result['매출액_2024'] = values[2]
                elif "영업이익" in row_name and "률" not in row_name:
                    result['영업이익_2024'] = values[2]
                elif "당기순이익" in row_name:
                    result['당기순이익_2024'] = values[2]
                elif "영업이익률" in row_name:
                    result['영업이익률_2024'] = values[2]
    return result

# Excel 작성 루프
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        ticker = cell.value
        if ticker and isinstance(ticker, str) and len(ticker) == 6 and ticker.isdigit():
            try:
                last_trading_day = get_last_trading_day(today, ticker)
                ohlcv = stock.get_market_ohlcv(last_trading_day, last_trading_day, ticker)
                price = f"{ohlcv['종가'].values[0]:,}"
                change_value = ohlcv['등락률'].values[0]
                change_text = f"{change_value:.2f}%"
                fundamental = stock.get_market_fundamental(last_trading_day, last_trading_day, ticker)
                per = fundamental['PER'].values[0]
                market_cap_data = stock.get_market_cap(last_trading_day, last_trading_day, ticker)
                market_cap_raw = int(market_cap_data['시가총액'].values[0])
                market_cap = format_market_cap(market_cap_raw)

                # pykrx 데이터 작성
                ws.cell(row=cell.row + 1, column=cell.column, value=price)
                change_cell = ws.cell(row=cell.row + 2, column=cell.column, value=change_text)
                if change_value > 0:
                    change_cell.font = Font(color="FF0000")
                elif change_value < 0:
                    change_cell.font = Font(color="0000FF")
                else:
                    change_cell.font = Font(color="000000")

                ws.cell(row=cell.row + 3, column=cell.column, value=market_cap)
                ws.cell(row=cell.row + 4, column=cell.column, value=per)

                # 재무제표 데이터 작성 (2024)
                financials = get_financials_2024(ticker)
                if financials:
                    ws.cell(row=cell.row + 6, column=cell.column, value=financials.get('매출액_2024', 'N/A'))
                    ws.cell(row=cell.row + 7, column=cell.column, value=financials.get('영업이익_2024', 'N/A'))
                    ws.cell(row=cell.row + 8, column=cell.column, value=financials.get('당기순이익_2024', 'N/A'))
                    ws.cell(row=cell.row + 9, column=cell.column, value=financials.get('영업이익률_2024', 'N/A'))

            except Exception as e:
                print(f"Error processing ticker {ticker}: {e}")

# 저장
wb.save(file_path)
print("✅ Update Complete")
